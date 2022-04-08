'''
Need to install these libraries:
pip install pillow
pip install openpyxl
pip install fpdf

Exaple of the required config.json:

{    
    "Input": {
        "pedirValores" : false
    },
    "Output": {
        "folderName" : "test",
        "generateSinglePDF" : true,
        "pdfPageSize" : "A4",
        "WidthMargin" : 20 ,
        "HighMargin" : 10,
        "porcentajeTamannoImagenEnPDF" : 100
    },
    "Imagen": {
        "nombreImagen" : "align.jpg",
        "PorcentajeAltNombre" : 43
    },
    "Excel": {
        "nombreArchivo" : "lista.xlsx",
        "nombreHoja" : "test",
        "nombreColumna" : "Preferred_Name"
    },
    "Letra": {
        "tamanno" : 36,
        "color": "black",
        "tipo" : "arial.ttf",
        "case" : "excel"
    }
}

'''
import openpyxl
import os
import os.path
from os import path
from os import replace
from pathlib import Path
import PIL
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
import json
from fpdf import FPDF

#Reading global properties from json file
with open('config.json') as f:
        data = json.load(f)
        # Input
        pedirValores = data['Input']['pedirValores']
        # Output
        outputfolderName = data['Output']['folderName']
        pdf_outputfolderName = data['Output']['pdf_folderName']
        generateSinglePDF = data['Output']['generateSinglePDF']
        pdfPageSize = data['Output']['pdfPageSize']
        w_margin = data['Output']['WidthMargin']
        h_margin = data['Output']['HighMargin']
        porcentajeTamannoImagenEnPDF = data['Output']['porcentajeTamannoImagenEnPDF']
        # imagen
        imagePath = data['Imagen']['nombreImagen']
        altNombre = data['Imagen']['PorcentajeAltNombre']
        # Excel
        csvPath = data['Excel']['nombreArchivo']
        nombreHoja = data['Excel']['nombreHoja']
        nombreCol = data['Excel']['nombreColumna']
        # Letra
        font_size = data['Letra']['tamanno']
        font_color= data['Letra']['color']
        font_type = data['Letra']['tipo']
        font_case = data['Letra']['case']

def create_folder(current_directory ,newdir):
    final_directory = os.path.join(current_directory, r'{}'.format(newdir))
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)

def get_params_from_user():
    print('  Capturando Parametros:')
    okflag = True

    while(okflag):
        global imagePath
        imagePath = input('    Cual es el nombre de la imagen para el diploma: ')
        if (path.exists(imagePath)):
            okflag = False
        else:
            print(f'      La imagen {imagePath} no existe')        
    
    okflag = True

    while(okflag):
        global altNombre
        altNombre = int(input('    Cual es el porcentaje de la altura (de arriba hacia abajo) a la que hay que colocar el nombre: '))
        if (altNombre>0 and altNombre<101):
            okflag = False
        else:
            print(f'      La altura debe ser entre 1 y 100')
    
    okflag = True
    
    while(okflag):
        global csvPath
        csvPath = input('    Cual es el nombre del csv con la lista de personas: ')
        if (path.exists(csvPath)):
            okflag = False
        else:
            print(f'     El archivo:  {csvPath} no existe')

    global nombreHoja
    nombreHoja = input('    Cual es el nombre de la hoja en el excel en donde estan los datos?: ')

    global nombreCol
    nombreCol = input('    Cual es el nombre de la columna donde estan los nombres en la hoja de excel: ')

    print('  Finalizando Captura de Parametros:')

def add_text_to_Image(personName, personNumber):

    #open image
    img = Image.open(imagePath)
    
    #Get image size
    #W,H = (829,649)
    W,H = img.size
    
    # Call draw Method to add 2D graphics in an image
    I1 = ImageDraw.Draw(img)
    
    # Custom font style and font size
    myFont = ImageFont.truetype(font_type, font_size)

    w, h = I1.textsize(personName,font=myFont)
    
    I1.text(((W-w)/2,(altNombre/100*H)), personName, font=myFont, fill = font_color)
    
    #img.show()

    img_path = f'./{outputfolderName}/{personName.replace(" ","_").replace(".","").strip()}.jpg'
    
    img.save(img_path)

    img.close()

    print(f'    Certificado #{personNumber} creado para: {personName}')

    # return Image w and h in mm
    dictionary ={
        "path": img_path,
        "w": W * 0.2645833333,
        "h": H * 0.2645833333
    }

    return (dictionary)

def read_names_from_file():
    print('  Iniciando Creacion de Certificados Individuales:')    
    
    personas = []

    xlsx_file = Path(csvPath)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    nombres = wb_obj[nombreHoja]

    # Determinar # columna donde esta el titulo especificado
    index = 0
    columnFound = False 
    #Recorre todas las columnas de la fila 1 y compara el header
    for column in nombres.iter_cols(1, nombres.max_column):
        if (column[0].value ==nombreCol):
            columnFound = True
            break
        index = index +1

    if (columnFound):
        # identificar el tipo de case en la letra que se va a usar (minuscula, mayuscula, titulo, la que tiene el exel)
        # Luego, recorre todas las filas de la columna que tiene el header especifico y guarda sus valores en un directorio
        
        if (font_case == 'title'):
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[index].value != None):
                    personas.append(row[index].value.title())
        elif (font_case == 'upper'):
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[index].value != None):
                    personas.append(row[index].value.upper())
        elif (font_case == 'lower'):
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[index].value != None):
                    personas.append(row[index].value.lower())
        else:
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[index].value != None):
                    personas.append(row[index].value)

        # si el excel tiene datos en esa columna se generan los certificados
        if (len(personas) >= 1):
            count = 1

            if (generateSinglePDF):
                W,H = (0,0)
                # set pdf page size accordig with provided parameter in config.json
                if (pdfPageSize == 'A4'):
                    W,H = (297,210)
                elif (pdfPageSize == 'Letter'):
                    W,H = (279.4, 215.9)
                elif (pdfPageSize == 'Legal'):
                    W,H = (355, 215)
                else: # default is A4
                    W,H = (297,210)
                
                pdf = FPDF('l','mm',(H,W))
                for persona in personas:
                    diploma = add_text_to_Image(persona, count)
                    count = count +1
                    pdf.add_page()
                    pdf.image(diploma['path'],
                            w_margin,
                            h_margin,
                            diploma['w']*(porcentajeTamannoImagenEnPDF/100),
                            diploma['h']*(porcentajeTamannoImagenEnPDF/100))

                pdf.output(f'{pdf_outputfolderName}/{outputfolderName}_Diplomas.pdf', "F")
                    
            else:
                for persona in personas:                    
                    add_text_to_Image(persona, count)
                    count = count +1

        else:
            print(f'    No existen valores en la columna: {nombreCol} en la hoja {nombreHoja} del archivo de excel.')
        
    else: 
        print(f'    La columna: {nombreCol} no existe en la hoja {nombreHoja} del archivo de excel.')

    print('  Finalizando Creacion de Certificados Individuales:')

def main():
    print('Iniciando programa para crear diplomas....')

    #create output Root Folder
    create_folder(os.getcwd(), outputfolderName)

    #create pdf output Root Folder
    create_folder(os.getcwd(), pdf_outputfolderName)

    if (pedirValores):
        get_params_from_user()

    read_names_from_file()
    
    print('--------------fin----------------------')

if __name__ == '__main__':
    main()

    #https://stackoverflow.com/questions/43767328/python-fpdf-not-sizing-correctly