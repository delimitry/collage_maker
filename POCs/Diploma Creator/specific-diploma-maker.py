'''
Need to install these libraries:
pip install pillow
pip install openpyxl
pip install fpdf

Este script lee un excel que tiene 2 columnas clave: Nombre Persona y tipo de certificado
Genera un pdf conteniendo todos los nombres del excel y generando un fondo (imagen) especifico para cada uno
basado en la columna tipos y su  valor. 

Exaple of the required config.json:

{    

    "Output": {
        "folderName" : "test",
        "pdf_folderName" : "PDFs",
        "generateSinglePDF" : true,
        "pdfPageSize" : "Letter",
        "WidthMargin" : 4,
        "HighMargin" : 6,
        "porcentajeTamannoImagenEnPDF" : 107
    },
    "Imagen": {
        "PorcentajeAltNombre" : 43
    },
    "Excel": {
        "nombreArchivo" : "lista.xlsx",
        "nombreHoja" : "testbigmama",
        "nombreColumna" : "Preferred_Name",
        "tiposColumna" : "Time",
        
        "imageType1_value" : "5 ANNOS",
        "imageType1_path" : "./letter/letter_align_5.JPG",
        
        "imageType2_value" : "10 ANNOS",
        "imageType2_path" : "./letter/letter_align_10.JPG",
                
        "imageType3_value" : "15 ANNOS",
        "imageType3_path" : "./letter/letter_align_15.JPG",
        
        "imageType4_value" : "20 ANNOS",
        "imageType4_path" : "./letter/letter_align_20.JPG"
    },
    "Letra": {
        "tamanno" : 36,
        "color": "#009BC7",
        "color-opcion2": "black",
        "tipo" : "arial.ttf",
        "case" : "excel"
    }
}

'''
import openpyxl
import os
import os.path
from os import path
from os import replace
from pathlib import Path
import PIL
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
import json
from fpdf import FPDF

#Reading global properties from json file
with open('specific-diploma-maker-config.json') as f:
        data = json.load(f)
        # Output
        outputfolderName = data['Output']['folderName']
        pdf_outputfolderName = data['Output']['pdf_folderName']
        generateSinglePDF = data['Output']['generateSinglePDF']
        pdfPageSize = data['Output']['pdfPageSize']
        w_margin = data['Output']['WidthMargin']
        h_margin = data['Output']['HighMargin']
        porcentajeTamannoImagenEnPDF = data['Output']['porcentajeTamannoImagenEnPDF']
        # imagen
        altNombre = data['Imagen']['PorcentajeAltNombre']
        # Excel
        csvPath = data['Excel']['nombreArchivo']
        nombreHoja = data['Excel']['nombreHoja']
        nombreCol = data['Excel']['nombreColumna']
        tiposCol = data['Excel']['tiposColumna']
        # Type 1 of picture 5 years
        imageType1_enabled = data['Excel']['imageType1_enabled']
        imageType1_value  = data['Excel']['imageType1_value']
        imageType1_path  = data['Excel']['imageType1_path']
        # Type 2 of picture 10 years
        imageType2_enabled = data['Excel']['imageType2_enabled']
        imageType2_value  = data['Excel']['imageType2_value']
        imageType2_path  = data['Excel']['imageType2_path']
        # Type 3 of picture 15 years
        imageType3_enabled = data['Excel']['imageType3_enabled']
        imageType3_value  = data['Excel']['imageType3_value']
        imageType3_path  = data['Excel']['imageType3_path']
        # Type 3 of picture 20 years
        imageType4_enabled = data['Excel']['imageType4_enabled']
        imageType4_value  = data['Excel']['imageType4_value']
        imageType4_path  = data['Excel']['imageType4_path']

        # Letra
        font_size = data['Letra']['tamanno']
        font_color= data['Letra']['color']
        font_type = data['Letra']['tipo']
        font_case = data['Letra']['case']

def create_folder(current_directory ,newdir):
    final_directory = os.path.join(current_directory, r'{}'.format(newdir))
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)

def add_text_to_Image(imagePath, personName, personNumber):

    #open image
    img = Image.open(imagePath)
    
    #Get image size
    #W,H = (829,649)
    W,H = img.size
    
    # Call draw Method to add 2D graphics in an image
    I1 = ImageDraw.Draw(img)
    
    # Custom font style and font size
    myFont = ImageFont.truetype(font_type, font_size)

    w, h = I1.textsize(personName,font=myFont)
    
    I1.text(((W-w)/2,(altNombre/100*H)), personName, font=myFont, fill = font_color)
    
    #img.show()

    img_path = f'./{outputfolderName}/{personName.replace(" ","_").replace(".","").strip()}.jpg'
    
    img.save(img_path)

    img.close()

    print(f'    Certificado #{personNumber} creado para: {personName}')

    # return Image w and h in mm
    dictionary ={
        "path": img_path,
        "w": W * 0.2645833333,
        "h": H * 0.2645833333
    }

    return (dictionary)

def transform_time_to_picture (time):
    ''' This function will map the cell value with the image path
        e.g 5 ANNOS with ./letter/letter_align_5.JPG
    '''
    picture = ""
    if (time == imageType1_value):
        picture = imageType1_path

    elif (time == imageType2_value):
        picture = imageType2_path

    elif (time == imageType3_value):
        picture = imageType3_path

    elif (time == imageType4_value):
        picture = imageType4_path

    return picture

def read_names_from_file():
    print('  Iniciando Creacion de Certificados Individuales:')    
    
    # lista de nombre de personas leidas del excel
    personas = []
    
    # lista de tipos de imagenes leidas del excel ( 5, 10, 15 etc)
    imageTypes = []

    xlsx_file = Path(csvPath)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    nombres = wb_obj[nombreHoja]

    # Determinar # columna donde esta el titulo especificado
    indexnombre = 0 # # columna de los nombres
    indextipo = 0 # # columna de los tipo de imagenes
    columnFound = False 

    #Recorre todas las columnas de la fila 1 y compara el header (buscar columna nombre)
    for column in nombres.iter_cols(1, nombres.max_column):
        if (column[0].value ==nombreCol):
            columnFound = True
            break
        indexnombre = indexnombre +1

    #Recorre todas las columnas de la fila 1 y compara el header (buscar columna annos)
    for column in nombres.iter_cols(1, nombres.max_column):
        if (column[0].value ==tiposCol):
            columnFound = True
            break
        indextipo = indextipo +1

    if (columnFound):
        # identificar el tipo de case en la letra que se va a usar (minuscula, mayuscula, titulo, la que tiene el exel)
        # Luego, recorre todas las filas de la columna que tiene el header especifico y guarda sus valores en un directorio
        
        # Cambia el font poniendo la primera letra de todas las palabras en mayuscula de lo que se lea en la celda
        if (font_case == 'title'):
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[indexnombre].value != None):
                    personas.append(row[indexnombre].value.title())
                    picture_path = transform_time_to_picture(row[indextipo].value)
                    imageTypes.append(picture_path)

        # Cambia el font a mayuscula de lo que se lea en la celda            
        elif (font_case == 'upper'):
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[indexnombre].value != None):
                    personas.append(row[indexnombre].value.upper())
                    picture_path = transform_time_to_picture(row[indextipo].value)
                    imageTypes.append(picture_path)
       # Cambia el font a minuscula de lo que se lea en la celda                         
        elif (font_case == 'lower'):
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[indexnombre].value != None):
                    personas.append(row[indexnombre].value.lower())
                    picture_path = transform_time_to_picture(row[indextipo].value)
                    imageTypes.append(picture_path)
        
        # no Cambia el font de la celda
        else:
            for row in nombres.iter_rows(2, nombres.max_row):
                if (row[indexnombre].value != None):
                    personas.append(row[indexnombre].value)
                    picture_path = transform_time_to_picture(row[indextipo].value)
                    imageTypes.append(picture_path)

        # si el excel tiene datos en esa columna se generan los certificados
        if (len(personas) >= 1):
            count = 1

            # si hay que generar un PDF.. entonces: 
            if (generateSinglePDF):
                W,H = (0,0)

                # set pdf page size accordig with provided parameter in config.json
                if (pdfPageSize == 'A4'):
                    W,H = (297,210)
                elif (pdfPageSize == 'Letter'):
                    W,H = (279.4, 215.9)
                elif (pdfPageSize == 'Legal'):
                    W,H = (355, 215)
                else: # default is A4
                    W,H = (297,210)
                
                pdf = FPDF('l','mm',(H,W))
                count2 = 0
                for persona in personas:
                    diploma = add_text_to_Image(imageTypes[count2], persona, count)
                    count = count +1
                    pdf.add_page()
                    pdf.image(diploma['path'],
                            w_margin,
                            h_margin,
                            diploma['w']*(porcentajeTamannoImagenEnPDF/100),
                            diploma['h']*(porcentajeTamannoImagenEnPDF/100))
                    count2 = count2 + 1

                pdf.output(f'{pdf_outputfolderName}/{outputfolderName}_Diplomas.pdf', "F")
                    
            else:
                for persona in personas:                    
                    add_text_to_Image(persona, count)
                    count = count +1

        else:
            print(f'    No existen valores en la columna: {nombreCol} en la hoja {nombreHoja} del archivo de excel.')
        
    else: 
        print(f'    La columna: {nombreCol} no existe en la hoja {nombreHoja} del archivo de excel.')

    print('  Finalizando Creacion de Certificados Individuales:')

def main():
    print('Iniciando programa para crear diplomas....')

    #create output Root Folder
    create_folder(os.getcwd(), outputfolderName)

    #create pdf output Root Folder
    create_folder(os.getcwd(), pdf_outputfolderName)

    read_names_from_file()
    
    
    print('--------------fin----------------------')

if __name__ == '__main__':
    main()

    #https://stackoverflow.com/questions/43767328/python-fpdf-not-sizing-correctly